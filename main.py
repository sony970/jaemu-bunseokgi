import argparse
import datetime
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import dart_client
from config import API_KEY

METRIC_ORDER = ["매출액", "영업이익", "당기순이익", "자산총계", "부채총계", "자본총계"]
OUTPUT_DIR = Path(__file__).parent / "output"


def fetch_company(api_key, corp_codes, name, year=None):
    corp = dart_client.find_corp(corp_codes, name)
    print(f"  [{name}] -> {corp['corp_name']} (corp_code={corp['corp_code']}, stock_code={corp['stock_code'] or '비상장'})")

    if year is not None:
        status, message, fs_div_used, accounts = dart_client.get_financial_accounts(api_key, corp["corp_code"], year)
        if status != "000" or not accounts:
            raise dart_client.DartApiError(
                f"'{corp['corp_name']}'의 {year}년 사업보고서를 찾을 수 없습니다 (status={status}, message={message})"
            )
        used_year = year
    else:
        current_year = datetime.date.today().year
        used_year, fs_div_used, accounts = dart_client.find_latest_available_year(
            api_key, corp["corp_code"], start_year=current_year - 1
        )
        if used_year is None:
            raise dart_client.DartApiError(f"'{corp['corp_name']}'의 최근 사업보고서를 찾을 수 없습니다.")

    print(f"    -> 사업연도 {used_year} 기준 ({'연결' if fs_div_used == 'CFS' else '별도'}재무제표) 조회 완료")

    metrics = dart_client.extract_key_metrics(accounts)
    years = {"당기": used_year, "전기": used_year - 1, "전전기": used_year - 2}

    return {
        "corp_name": corp["corp_name"],
        "used_year": used_year,
        "fs_div": fs_div_used,
        "years": years,
        "metrics": metrics,
        "raw_accounts": accounts,
    }


def build_summary_df(companies):
    columns = []
    data = {}
    for c in companies:
        for period in ("당기", "전기", "전전기"):
            col = (c["corp_name"], f"{c['years'][period]}년")
            columns.append(col)
            data[col] = [c["metrics"][m][period] for m in METRIC_ORDER]

    df = pd.DataFrame(data, index=METRIC_ORDER)
    df.columns = pd.MultiIndex.from_tuples(df.columns, names=["회사", "연도"])
    return df


def build_detail_df(company):
    rows = []
    for acc in company["raw_accounts"]:
        rows.append(
            {
                "구분": acc.get("sj_nm"),
                "계정명": acc.get("account_nm"),
                f"당기({company['years']['당기']})": acc.get("thstrm_amount"),
                f"전기({company['years']['전기']})": acc.get("frmtrm_amount"),
                f"전전기({company['years']['전전기']})": acc.get("bfefrmtrm_amount"),
            }
        )
    return pd.DataFrame(rows)


def autosize_columns(worksheet, min_width=10, max_width=40):
    for col_cells in worksheet.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def save_excel(companies, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = build_summary_df(companies)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="비교요약")
        ws = writer.sheets["비교요약"]
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        autosize_columns(ws)

        for c in companies:
            sheet_name = c["corp_name"][:31]
            detail_df = build_detail_df(c)
            detail_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws2 = writer.sheets[sheet_name]
            for cell in ws2[1]:
                cell.font = Font(bold=True)
            autosize_columns(ws2)

    print(f"\n엑셀 파일 저장 완료: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="DART 기반 3개 회사 재무제표 비교기")
    parser.add_argument("companies", nargs=3, metavar=("회사1", "회사2", "회사3"), help="비교할 회사명 3개")
    parser.add_argument("--year", type=int, default=None, help="기준 사업연도 (예: 2024). 생략 시 최신 연도 자동 탐색")
    parser.add_argument("--output", type=str, default=None, help="저장할 엑셀 파일 경로")
    args = parser.parse_args()

    print("DART에서 회사 고유번호 목록을 불러오는 중...")
    corp_codes = dart_client.load_corp_codes(API_KEY)
    print(f"고유번호 {len(corp_codes)}건 로드 완료\n")

    companies = []
    for name in args.companies:
        try:
            company = fetch_company(API_KEY, corp_codes, name, year=args.year)
            companies.append(company)
        except dart_client.CompanyNotFoundError as e:
            print(f"오류: {e}")
            sys.exit(1)
        except dart_client.DartApiError as e:
            print(f"오류: {e}")
            sys.exit(1)

    if args.output:
        output_path = Path(args.output)
    else:
        today = datetime.date.today().strftime("%Y%m%d")
        output_path = OUTPUT_DIR / f"재무비교_{today}.xlsx"

    save_excel(companies, output_path)


if __name__ == "__main__":
    main()
